"""
Forward-Looking WBR Report ETL Pipeline

Automates weekly forward-looking WBR report from case review queue data.
Reads multi-sheet Excel workbook, applies queue/category filters to create
filtered views, computes statistical metrics (Average, TP90, TP99),
age-group distributions, and average processing time differences.

Preserves source formatting in output workbook.

Saves ~2 hours/week of manual Excel filtering and stats computation.
"""

import pandas as pd
import numpy as np
import glob
import os
from copy import copy
from openpyxl import Workbook, load_workbook

# --- Configuration ---
INPUT_DIR = os.environ.get("WBR_DATA_DIR", os.path.join(".", "data"))
OUTPUT_FILE = os.path.join(INPUT_DIR, "FW_WBR_Output.xlsx")

HV_QUEUE = "queue-hv-review@example.com"
LV_QUEUE = "queue-lv-review@example.com"
HV_CATEGORY = "HV Investigation Review - Regional"
LV_CATEGORY = "LV Investigation Review - Regional"

STATS_COLS_LOWER = ["transfer to review", "in review", "time to close"]
STATS_LABELS = ["Transfer to Review", "In Review", "Time to Close"]

SHEET_CONFIG = {
    "Unassigned Cases": [
        ("Unassigned Cases", None, None),
        ("Unassigned Cases (HV)", "Queue", HV_QUEUE),
        ("Unassigned Cases (LV)", "Queue", LV_QUEUE),
    ],
    "Unresolved Cases": [
        ("Unresolved Cases", None, None),
        ("Unresolved Cases (HV)", "Queue", HV_QUEUE),
        ("Unresolved Cases (LV)", "Queue", LV_QUEUE),
    ],
    "Resolved Cases": [
        ("Resolved Cases", None, None),
        ("Resolved Cases (HV)", "Category", HV_CATEGORY),
        ("Resolved Cases (LV)", "Category", LV_CATEGORY),
    ],
}


def get_latest_file(directory, prefix):
    """Find the most recently modified file matching a prefix pattern."""
    pattern = os.path.join(directory, f"{prefix}*.xlsx")
    files = glob.glob(pattern)
    if not files:
        raise FileNotFoundError(f"No files matching '{prefix}*.xlsx' in {directory}")
    return max(files, key=os.path.getmtime)


def find_col_df(df, name):
    """Case-insensitive column name lookup."""
    target = name.strip().lower()
    for c in df.columns:
        if isinstance(c, str) and c.strip().lower() == target:
            return c
    return None


def find_col_contains_df(df, *keywords):
    """Find column whose name contains all given keywords (case-insensitive)."""
    for c in df.columns:
        if not isinstance(c, str):
            continue
        if all(kw in c.strip().lower() for kw in keywords):
            return c
    return None


def find_stats_area(ws):
    """Locate the stats table (Average/TP90/TP99) in a worksheet."""
    header_row = None
    header_col = None
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            if isinstance(val, str) and val.strip().lower() == "average":
                next_val = ws.cell(row=row, column=col + 1).value
                if isinstance(next_val, str) and "tp" in next_val.strip().lower():
                    header_row, header_col = row, col
                    break
        if header_row:
            break

    if header_row is None:
        return None

    label_col = header_col - 1
    end_row = header_row
    for r in range(header_row + 1, header_row + 10):
        val = ws.cell(row=r, column=label_col).value
        if val is not None and isinstance(val, str) and val.strip() != "":
            end_row = r
        else:
            break

    return {
        "header_row": header_row,
        "label_col": label_col,
        "end_row": end_row,
        "end_col": header_col + 2,
    }


def copy_sheet_full(source_wb, source_sheet_name, target_wb, target_sheet_name):
    """Full copy of a sheet preserving values and formatting."""
    src_ws = source_wb[source_sheet_name]
    tgt_ws = target_wb.create_sheet(title=target_sheet_name)
    for row in src_ws.iter_rows(min_row=1, max_row=src_ws.max_row, max_col=src_ws.max_column):
        for cell in row:
            new_cell = tgt_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.alignment = copy(cell.alignment)
    for col_letter, dim in src_ws.column_dimensions.items():
        tgt_ws.column_dimensions[col_letter].width = dim.width
    for merged_range in src_ws.merged_cells.ranges:
        tgt_ws.merge_cells(str(merged_range))
    return tgt_ws


def compute_stats_from_df(df):
    """Compute [average, tp90, tp99] for each stats column."""
    results = []
    for col_name in STATS_COLS_LOWER:
        col = find_col_df(df, col_name)
        if col is None:
            results.append([0, 0, 0])
            continue
        vals = pd.to_numeric(df[col], errors="coerce").dropna()
        if len(vals) == 0:
            results.append([0, 0, 0])
        else:
            results.append([
                round(float(vals.mean()), 2),
                round(float(np.percentile(vals, 90)), 2),
                round(float(np.percentile(vals, 99)), 2),
            ])
    return results


def build_filtered_sheet(target_wb, sheet_name, filtered_df, source_ws, stats_area):
    """Build a filtered sheet with recalculated stats, preserving source formatting."""
    ws = target_wb.create_sheet(title=sheet_name)

    # Write headers with formatting
    for col_idx, col_name in enumerate(filtered_df.columns, 1):
        src_cell = source_ws.cell(row=1, column=col_idx)
        new_cell = ws.cell(row=1, column=col_idx, value=col_name)
        if src_cell.has_style:
            new_cell.font = copy(src_cell.font)
            new_cell.border = copy(src_cell.border)
            new_cell.fill = copy(src_cell.fill)
            new_cell.number_format = src_cell.number_format
            new_cell.alignment = copy(src_cell.alignment)

    # Write data rows
    for row_idx, (_, row_data) in enumerate(filtered_df.iterrows(), 2):
        for col_idx, val in enumerate(row_data, 1):
            if pd.notna(val):
                ws.cell(row=row_idx, column=col_idx, value=val)

    # Copy column widths
    for col_letter, dim in source_ws.column_dimensions.items():
        ws.column_dimensions[col_letter].width = dim.width

    # Recreate stats table with recalculated values
    if stats_area:
        label_col = stats_area["label_col"]
        header_col = label_col + 1
        tgt_header_row = stats_area["header_row"]

        for col in range(header_col, stats_area["end_col"] + 1):
            src_cell = source_ws.cell(row=tgt_header_row, column=col)
            new_cell = ws.cell(row=tgt_header_row, column=col, value=src_cell.value)
            if src_cell.has_style:
                new_cell.font = copy(src_cell.font)
                new_cell.border = copy(src_cell.border)
                new_cell.fill = copy(src_cell.fill)
                new_cell.alignment = copy(src_cell.alignment)

        stats_values = compute_stats_from_df(filtered_df)
        for i, (label, vals) in enumerate(zip(STATS_LABELS, stats_values)):
            tgt_row = tgt_header_row + 1 + i
            src_row = tgt_header_row + 1 + i

            src_cell = source_ws.cell(row=src_row, column=label_col)
            new_cell = ws.cell(row=tgt_row, column=label_col, value=label)
            if src_cell.has_style:
                new_cell.font = copy(src_cell.font)
                new_cell.border = copy(src_cell.border)
                new_cell.fill = copy(src_cell.fill)
                new_cell.alignment = copy(src_cell.alignment)

            for j, v in enumerate(vals):
                src_cell = source_ws.cell(row=src_row, column=header_col + j)
                new_cell = ws.cell(row=tgt_row, column=header_col + j, value=v)
                if src_cell.has_style:
                    new_cell.font = copy(src_cell.font)
                    new_cell.border = copy(src_cell.border)
                    new_cell.fill = copy(src_cell.fill)
                    new_cell.number_format = src_cell.number_format
                    new_cell.alignment = copy(src_cell.alignment)

    return ws


def compute_avg_diff_days(df):
    """Compute average difference in days between transfer and creation dates."""
    transfer_col = find_col_contains_df(df, "transfer_date")
    create_col = find_col_contains_df(df, "create_date")
    if not transfer_col or not create_col:
        return 0
    transfer_dt = pd.to_datetime(df[transfer_col], errors="coerce")
    create_dt = pd.to_datetime(df[create_col], errors="coerce")
    diffs = ((transfer_dt - create_dt).dt.total_seconds() / 86400).dropna()
    return round(float(diffs.mean()), 2) if len(diffs) > 0 else 0


def compute_age_pivot(df):
    """Build age-group distribution (0-10, 10-20, etc.) for review time."""
    review_col = find_col_df(df, "in review")
    if review_col is None:
        return [["<0 or (blank)", ""], ["Grand Total", 0]]

    raw = pd.to_numeric(df[review_col], errors="coerce")
    blank_count = int(raw.isna().sum())
    vals = raw.dropna()
    neg_count = int((vals < 0).sum()) if len(vals) > 0 else 0
    pos_vals = vals[vals >= 0]

    bins = {}
    for v in pos_vals:
        bucket = int(v // 10) * 10
        bins[bucket] = bins.get(bucket, 0) + 1

    rows = []
    combined = neg_count + blank_count
    rows.append(["<0 or (blank)", combined if combined > 0 else ""])
    for bucket in sorted(bins.keys()):
        rows.append([f"{bucket}-{bucket + 10}", bins[bucket]])
    rows.append(["Grand Total", len(vals) + blank_count])
    return rows


def find_true_max_col(ws):
    """Find the rightmost column with actual data."""
    max_col = 1
    for row in range(1, ws.max_row + 1):
        for col in range(ws.max_column, 0, -1):
            val = ws.cell(row=row, column=col).value
            if val is not None and str(val).strip() != "":
                max_col = max(max_col, col)
                break
    return max_col


def find_stats_row_in_output(ws):
    """Find the row containing stats labels in the output sheet."""
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            if isinstance(val, str) and val.strip().lower() == STATS_LABELS[0].lower():
                return row
    return None


def write_extras(ws, avg_diff, pivot):
    """Write Avg Diff Days and age-group pivot to the right of existing content."""
    safe_col = find_true_max_col(ws) + 2
    stats_row = find_stats_row_in_output(ws)
    anchor_row = stats_row if stats_row else 2

    ws.cell(row=anchor_row, column=safe_col, value="Avg Processing Days")
    ws.cell(row=anchor_row + 1, column=safe_col, value=avg_diff)

    pivot_col = safe_col + 2
    ws.cell(row=anchor_row, column=pivot_col, value="Age Group")
    ws.cell(row=anchor_row, column=pivot_col + 1, value="Count")
    for i, prow in enumerate(pivot):
        ws.cell(row=anchor_row + 1 + i, column=pivot_col, value=prow[0])
        ws.cell(row=anchor_row + 1 + i, column=pivot_col + 1, value=prow[1])


def run_pipeline():
    print("=" * 60)
    print("  Forward-Looking WBR ETL Pipeline")
    print("=" * 60)

    print("\n[1/5] Loading source file...")
    source_file = get_latest_file(INPUT_DIR, "review_queue_")
    print(f"  File: {os.path.basename(source_file)}")

    source_wb = load_workbook(source_file)

    for sheet_name in SHEET_CONFIG:
        if sheet_name not in source_wb.sheetnames:
            print(f"  ERROR: Sheet '{sheet_name}' not found")
            return
        print(f"  Found '{sheet_name}'")

    source_dfs = {}
    stats_areas = {}
    for sheet_name in SHEET_CONFIG:
        source_dfs[sheet_name] = pd.read_excel(source_file, sheet_name=sheet_name)
        source_dfs[sheet_name].columns = [
            c.strip() if isinstance(c, str) else c
            for c in source_dfs[sheet_name].columns
        ]
        stats_areas[sheet_name] = find_stats_area(source_wb[sheet_name])
        print(f"  Loaded '{sheet_name}': {len(source_dfs[sheet_name])} rows")

    print("\n[2/5] Creating output workbook...")
    target_wb = Workbook()
    target_wb.remove(target_wb.active)

    print("\n[3/5] Building sheets...")
    output_sheets = {}
    for source_name, variants in SHEET_CONFIG.items():
        source_ws = source_wb[source_name]
        df_full = source_dfs[source_name]
        stats_area = stats_areas[source_name]

        for out_name, filter_col, filter_val in variants:
            if filter_col is None:
                ws = copy_sheet_full(source_wb, source_name, target_wb, out_name)
            else:
                col = find_col_df(df_full, filter_col)
                if col:
                    filtered_df = df_full[df_full[col].astype(str).str.strip() == filter_val].copy()
                else:
                    filtered_df = df_full.copy()
                print(f"  '{out_name}' -> filtered ({len(df_full)} -> {len(filtered_df)} rows)")
                ws = build_filtered_sheet(target_wb, out_name, filtered_df, source_ws, stats_area)

            output_sheets[out_name] = (ws, filter_col, filter_val)

    print("\n[4/5] Adding processing metrics and age distributions...")
    for source_name, variants in SHEET_CONFIG.items():
        df_full = source_dfs[source_name]
        for out_name, filter_col, filter_val in variants:
            ws = output_sheets[out_name][0]

            if filter_col:
                col = find_col_df(df_full, filter_col)
                filtered_df = df_full[df_full[col].astype(str).str.strip() == filter_val].copy() if col else df_full.copy()
            else:
                filtered_df = df_full

            avg_diff = compute_avg_diff_days(filtered_df)
            pivot = compute_age_pivot(filtered_df)
            print(f"  '{out_name}': Avg Days={avg_diff}, Age groups={len(pivot)-1}")
            write_extras(ws, avg_diff, pivot)

    print(f"\n[5/5] Saving output to {OUTPUT_FILE}...")
    target_wb.save(OUTPUT_FILE)

    print("\n" + "=" * 60)
    print("Done! Output sheets:")
    for source_name, variants in SHEET_CONFIG.items():
        for out_name, _, _ in variants:
            print(f"  - {out_name}")
    print("=" * 60)


if __name__ == "__main__":
    run_pipeline()
